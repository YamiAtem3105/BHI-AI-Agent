"""Convert file Excel cá nhân (vd: 'Quản lý công việc - Phan Minh Hoàng.xlsx')
   → JSON đặt trong data/personal/<slug>.json

JSON schema:
{
  "user": "Phan Minh Hoàng",
  "tasks": [ {id, name, start_date, end_date, duration, predecessor, status,
              elapsed, pic, support, reviewer, note, hours, log_history}, ... ],
  "archive": [ {report_date, user, role, project, content, hours, status,
                deadline, progress, note}, ... ]
}

Cách chạy: python scripts/convert_personal_xlsx.py
"""
import json
import os
import sys
import unicodedata
from datetime import datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(ROOT, "out s req")
OUT_DIR = os.path.join(ROOT, "data", "personal")

# Map file → user
FILE_TO_USER = {
    "Quản lý công việc - Phan Minh Hoàng.xlsx": "Phan Minh Hoàng",
}


def slugify(name: str) -> str:
    """'Phan Minh Hoàng' → 'phan_minh_hoang'"""
    nfkd = unicodedata.normalize("NFKD", name)
    no_accent = "".join(c for c in nfkd if not unicodedata.combining(c))
    no_accent = no_accent.replace("đ", "d").replace("Đ", "d")
    return no_accent.lower().replace(" ", "_")


def fmt_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def fmt_datetime(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v)


def num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def s(v):
    return str(v).strip() if v is not None else ""


def extract_tasks(ws) -> list:
    """Tab 'KẾ HOẠCH CHUYỂN ĐỔI SỐ' cá nhân — header ở R2, data từ R3.
    Cols: ID, Danh sách công việc, Ngày bắt đầu, Ngày kết thúc, Thời lượng,
          Việc trước, Trạng thái, Thời gian đã qua, Thực hiện, Phối Hợp,
          Kiểm tra, Ghi chú, Số giờ làm, Lịch sử ghi chú
    """
    tasks = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not row[0]:
            continue
        tasks.append({
            "id": s(row[0]),
            "name": s(row[1]),
            "start_date": fmt_date(row[2]),
            "end_date": fmt_date(row[3]),
            "duration_days": num(row[4]),
            "predecessor": s(row[5]),
            "status": s(row[6]),
            "elapsed": num(row[7]),
            "pic": s(row[8]),
            "support": s(row[9]),
            "reviewer": s(row[10]),
            "note": s(row[11]),
            "hours": num(row[12]) if len(row) > 12 else None,
            "log_history": s(row[13]) if len(row) > 13 else "",
        })
    return tasks


def extract_archive(ws) -> list:
    """Tab '_ARCHIVE_LOG' cá nhân — header R1, data từ R2.
    Cols: Ngày báo cáo, User, Vai trò, Dự án, Nội dung, Thời lượng,
          Trạng thái, Deadline, Tiến độ, Ghi chú
    """
    archive = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        archive.append({
            "report_date": fmt_datetime(row[0]),
            "user": s(row[1]),
            "role": s(row[2]),
            "project": s(row[3]),
            "task_content": s(row[4]),
            "hours": num(row[5]),
            "status": s(row[6]),
            "deadline": fmt_date(row[7]),
            "progress": s(row[8]) if len(row) > 8 else "",
            "report_note": s(row[9]) if len(row) > 9 else "",
        })
    return archive


def convert_file(file_path: str, user: str) -> dict:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    tasks, archive = [], []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "KẾ HOẠCH CHUYỂN ĐỔI SỐ":
            tasks = extract_tasks(ws)
        elif sheet_name == "_ARCHIVE_LOG":
            archive = extract_archive(ws)
    wb.close()
    return {"user": user, "tasks": tasks, "archive": archive}


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    summary = []
    for fname, user in FILE_TO_USER.items():
        src = os.path.join(SOURCE_DIR, fname)
        if not os.path.exists(src):
            print(f"  SKIP (không thấy): {src}")
            continue
        data = convert_file(src, user)
        out_path = os.path.join(OUT_DIR, slugify(user) + ".json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        summary.append({
            "user": user,
            "file": out_path,
            "tasks": len(data["tasks"]),
            "archive": len(data["archive"]),
        })
        print(f"  OK {user}: {len(data['tasks'])} tasks + {len(data['archive'])} archive → {out_path}")

    print("\nTổng kết:")
    for s_ in summary:
        print(f"  - {s_['user']}: tasks={s_['tasks']}, archive={s_['archive']}")


if __name__ == "__main__":
    main()
